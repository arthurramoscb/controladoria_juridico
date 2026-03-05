# Databricks notebook source
# MAGIC %run /Workspace/Jurídico/funcao_tratamento_fechamento/common_functions

# COMMAND ----------

from openpyxl import load_workbook

# diretorio projeto
diretorio_projeto = '/Volumes/databox/juridico_comum/arquivos/trabalhista/base_semanal_acordos/divisao_escritorios/'

arquivo_padrao = f'{diretorio_projeto}MODELO REPORTE SEMANAL.xlsx'

path_report_semanal = f'{diretorio_projeto}MODELO REPORTE SEMANAL_JUNHO_24 1.xlsx'

df_report_semanal = read_excel(path_report_semanal,"'BASE APTA'!A2:AJ1048576")

# # Especifique o caminho do arquivo Excel existente
# file_path = "/dbfs/path/to/your/existing_excel_file.xlsx"

# # Carregue o arquivo Excel
# wb = load_workbook(file_path, keep_vba=True)

# COMMAND ----------

import re
from openpyxl import load_workbook
import time
import pandas as pd

df_report_semanal = df_report_semanal.fillna('')

lista_negociadores = df_report_semanal.select('NEGOCIADOR').distinct().collect()

for x, negociador_ in enumerate(lista_negociadores):
    negociador_str = str(negociador_)
    negociador_xa = re.sub(r'xa0', ' ', negociador_str)
    negociador2 = re.sub(r'\\',' ', negociador_xa)
    negociador_for_split = re.sub(r'\s{2,}',' ',negociador2)
    negociador_split = negociador_for_split.split('=')
    negociador_txt = negociador_split[1]
    negociador = negociador_txt[1:len(negociador_txt)-2]

    df_negociador_ = df_report_semanal.filter(df_report_semanal['NEGOCIADOR'] == negociador)

    df_negociador_pandas =  df_negociador_.toPandas()

    path_destino = f'{diretorio_projeto}{negociador}.xlsx'

    path_temporario = f'/local_disk0/tmp/{negociador}.xlsx'

    if x == 0:
        dbutils.fs.cp(arquivo_padrao, path_destino, True)
        time.sleep(5)

        wb = load_workbook(path_destino, keep_vba=True, keep_links=True)

        # Suponha que você queira trabalhar na primeira planilha (índice 0)
        # print(wb.sheetnames)

        sheet = wb.worksheets[0]

        # # Determine onde começar a escrever os dados (por exemplo, a partir da célula A1)
        start_row = 2
        start_col = 1

        # # Escreva cabeçalhos (nomes das colunas) na primeira linha
        # for c_idx, header in enumerate(df_negociador_pandas.columns):
        #     sheet.cell(row=start_row, column=start_col + c_idx, value=header)

        # # Itere sobre as linhas do DataFrame e escreva nas células correspondentes na planilha Excel
        for r_idx, row in df_negociador_pandas.iterrows():
            for c_idx, value in enumerate(row):
                sheet.cell(row=start_row + 1 + r_idx, column=start_col + c_idx, value=value)

        wb.save(path_destino)
        # display(df_negociador_)

        wb.close()

        # dbutils.fs.cp(path_temporario, path_destino, True)

    

# COMMAND ----------

dbutils.fs.ls('/local_disk0/tmp')

# COMMAND ----------

wb.close()

# COMMAND ----------

print('')
